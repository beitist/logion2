import os
import re
import html
import shutil
from collections import defaultdict
import openpyxl
from app.logger import get_logger

logger = get_logger("ExcelAssembler")


def reassemble_xlsx(input_path: str, output_path: str, segments) -> None:
    """
    Reassembles an XLSX file by injecting translated content into the original cells.

    Each segment carries metadata with sheet, row, col, and sentence_index.
    Segments from the same cell are joined back together (space-separated)
    in sentence_index order to reconstruct the full cell translation.
    """
    shutil.copy(input_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    # Group segments by (sheet, row, col)
    cell_groups = defaultdict(list)
    for seg in segments:
        meta = seg.metadata or {}
        sheet = meta.get("sheet")
        row = meta.get("row")
        col = meta.get("col")
        if sheet is None or row is None or col is None:
            logger.warning(f"Segment {seg.segment_id} missing cell metadata, skipping")
            continue
        cell_groups[(sheet, row, col)].append(seg)

    # Sort each group by sentence_index and join
    for (sheet_name, row, col), segs in cell_groups.items():
        segs.sort(key=lambda s: (s.metadata or {}).get("sentence_index", 0))

        # Use target_content if available, otherwise source_text
        parts = []
        for s in segs:
            text = s.target_content if s.target_content else s.source_text
            # Strip any remaining HTML tags from tiptap, then decode entities
            text = re.sub(r'<[^>]+>', '', text)
            text = html.unescape(text)
            parts.append(text.strip())

        cell_text = " ".join(parts)

        # Write into the workbook
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.cell(row=row, column=col, value=cell_text)
        else:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook, skipping")

    wb.save(output_path)
    logger.info(f"XLSX reassembly complete: {output_path}")
